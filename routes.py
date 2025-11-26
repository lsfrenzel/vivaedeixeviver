from flask import render_template, request, redirect, url_for, flash, jsonify, abort
from flask_login import login_user, logout_user, login_required, current_user
from app import app, db, login_manager
from models import Voluntario, Hospital, Livro, Diario
from datetime import datetime, timedelta
from sqlalchemy import func, extract
from functools import wraps


@login_manager.user_loader
def load_user(user_id):
    return Voluntario.query.get(int(user_id))


@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.is_admin:
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        if current_user.is_admin:
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        email = request.form.get('email')
        senha = request.form.get('senha')
        
        voluntario = Voluntario.query.filter_by(email=email).first()
        
        if voluntario and voluntario.check_senha(senha):
            login_user(voluntario)
            if voluntario.is_admin:
                return redirect(url_for('admin_dashboard'))
            return redirect(url_for('dashboard'))
        else:
            flash('E-mail ou senha inválidos.', 'error')
    
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    hoje = datetime.now()
    inicio_mes = hoje.replace(day=1)
    inicio_ano = hoje.replace(month=1, day=1)
    
    horas_mes = db.session.query(func.sum(Diario.duracao)).filter(
        Diario.voluntario_id == current_user.id,
        Diario.data >= inicio_mes
    ).scalar() or 0
    
    horas_ano = db.session.query(func.sum(Diario.duracao)).filter(
        Diario.voluntario_id == current_user.id,
        Diario.data >= inicio_ano
    ).scalar() or 0
    
    total_atuacoes = Diario.query.filter_by(voluntario_id=current_user.id).count()
    
    recentes = Diario.query.filter_by(voluntario_id=current_user.id).order_by(
        Diario.data.desc()
    ).limit(5).all()
    
    horas_por_mes = []
    for mes in range(1, 13):
        total = db.session.query(func.sum(Diario.duracao)).filter(
            Diario.voluntario_id == current_user.id,
            extract('year', Diario.data) == hoje.year,
            extract('month', Diario.data) == mes
        ).scalar() or 0
        horas_por_mes.append(float(total))
    
    medalha = None
    if horas_mes >= 20:
        medalha = {'tipo': 'ouro', 'titulo': 'Voluntário Ouro', 'descricao': f'{horas_mes:.1f} horas este mês!'}
    elif horas_mes >= 10:
        medalha = {'tipo': 'prata', 'titulo': 'Voluntário Prata', 'descricao': f'{horas_mes:.1f} horas este mês!'}
    elif horas_mes >= 5:
        medalha = {'tipo': 'bronze', 'titulo': 'Voluntário Bronze', 'descricao': f'{horas_mes:.1f} horas este mês!'}
    
    return render_template('dashboard.html',
                         horas_mes=horas_mes,
                         horas_ano=horas_ano,
                         total_atuacoes=total_atuacoes,
                         recentes=recentes,
                         horas_por_mes=horas_por_mes,
                         medalha=medalha,
                         ano_atual=hoje.year)


@app.route('/nova-atuacao', methods=['GET', 'POST'])
@login_required
def nova_atuacao():
    if request.method == 'POST':
        data = request.form.get('data')
        periodo = request.form.get('periodo')
        duracao = request.form.get('duracao')
        relato = request.form.get('relato_qualitativo')
        
        faixas_etarias = ['0-3', '4-6', '7-9', '10-12', '13-15', '16-18']
        pacientes = {}
        for faixa in faixas_etarias:
            fem = request.form.get(f'pacientes_{faixa}_f') or '0'
            masc = request.form.get(f'pacientes_{faixa}_m') or '0'
            pacientes[faixa] = {'feminino': int(fem), 'masculino': int(masc)}
        
        locais = request.form.getlist('locais_atendimento')
        
        livros_ids = request.form.getlist('livros_selecionados')
        livros_data = []
        for livro_id in livros_ids:
            if livro_id:
                livro = Livro.query.get(int(livro_id))
                if livro:
                    livros_data.append({
                        'id': livro.id,
                        'titulo': livro.titulo,
                        'autor': livro.autor,
                        'editora': livro.editora
                    })
        
        novo_livro_titulo = request.form.get('novo_livro_titulo')
        if novo_livro_titulo:
            novo_livro_autor = request.form.get('novo_livro_autor')
            novo_livro_editora = request.form.get('novo_livro_editora')
            
            livro_existente = Livro.query.filter_by(titulo=novo_livro_titulo).first()
            if not livro_existente:
                novo_livro = Livro(
                    titulo=novo_livro_titulo,
                    autor=novo_livro_autor,
                    editora=novo_livro_editora
                )
                db.session.add(novo_livro)
                db.session.flush()
                livros_data.append({
                    'id': novo_livro.id,
                    'titulo': novo_livro.titulo,
                    'autor': novo_livro.autor,
                    'editora': novo_livro.editora
                })
        
        diario = Diario(
            voluntario_id=current_user.id,
            data=datetime.strptime(data, '%Y-%m-%d').date(),
            periodo=periodo,
            duracao=float(duracao),
            pacientes_atendidos=pacientes,
            locais_atendimento=locais,
            livros_contados=livros_data,
            relato_qualitativo=relato
        )
        
        db.session.add(diario)
        db.session.commit()
        
        flash('Atuação registrada com sucesso!', 'success')
        return redirect(url_for('confirmacao', diario_id=diario.id))
    
    locais_disponiveis = [
        'Leito',
        'UTI',
        'Pronto Socorro',
        'Hemodiálise',
        'Oncologia',
        'Pediatria',
        'Enfermaria',
        'Sala de Espera'
    ]
    
    data_atual = datetime.now().strftime('%Y-%m-%d')
    
    return render_template('nova_atuacao.html',
                         locais_disponiveis=locais_disponiveis,
                         data_atual=data_atual)


@app.route('/confirmacao/<int:diario_id>')
@login_required
def confirmacao(diario_id):
    diario = Diario.query.get_or_404(diario_id)
    
    if diario.voluntario_id != current_user.id:
        flash('Acesso negado.', 'error')
        return redirect(url_for('dashboard'))
    
    total_pacientes = 0
    for faixa, dados in diario.pacientes_atendidos.items():
        total_pacientes += dados.get('feminino', 0) + dados.get('masculino', 0)
    
    return render_template('confirmacao.html', diario=diario, total_pacientes=total_pacientes)


@app.route('/api/buscar-livros')
@login_required
def buscar_livros():
    termo = request.args.get('q', '')
    
    if len(termo) < 2:
        return jsonify([])
    
    livros = Livro.query.filter(
        Livro.titulo.ilike(f'%{termo}%')
    ).limit(10).all()
    
    resultados = [{
        'id': livro.id,
        'titulo': livro.titulo,
        'autor': livro.autor,
        'editora': livro.editora
    } for livro in livros]
    
    return jsonify(resultados)


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('Acesso negado. Apenas administradores podem acessar esta página.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function


@app.route('/admin')
@login_required
@admin_required
def admin_dashboard():
    hoje = datetime.now()
    inicio_mes = hoje.replace(day=1)
    inicio_ano = hoje.replace(month=1, day=1)
    
    total_voluntarios = Voluntario.query.count()
    total_atuacoes = Diario.query.count()
    
    horas_mes_total = db.session.query(func.sum(Diario.duracao)).filter(
        Diario.data >= inicio_mes
    ).scalar() or 0
    
    horas_ano_total = db.session.query(func.sum(Diario.duracao)).filter(
        Diario.data >= inicio_ano
    ).scalar() or 0
    
    total_pacientes = 0
    diarios = Diario.query.all()
    for diario in diarios:
        if diario.pacientes_atendidos:
            for faixa, dados in diario.pacientes_atendidos.items():
                total_pacientes += dados.get('feminino', 0) + dados.get('masculino', 0)
    
    total_livros = Livro.query.count()
    
    voluntarios_ativos = db.session.query(
        Voluntario.nome,
        func.sum(Diario.duracao).label('total_horas')
    ).join(Diario).filter(
        Diario.data >= inicio_mes
    ).group_by(Voluntario.id, Voluntario.nome).order_by(
        func.sum(Diario.duracao).desc()
    ).limit(10).all()
    
    horas_por_mes = []
    for mes in range(1, 13):
        total = db.session.query(func.sum(Diario.duracao)).filter(
            extract('year', Diario.data) == hoje.year,
            extract('month', Diario.data) == mes
        ).scalar() or 0
        horas_por_mes.append(float(total))
    
    return render_template('admin_dashboard.html',
                         total_voluntarios=total_voluntarios,
                         total_atuacoes=total_atuacoes,
                         horas_mes_total=horas_mes_total,
                         horas_ano_total=horas_ano_total,
                         total_pacientes=total_pacientes,
                         total_livros=total_livros,
                         voluntarios_ativos=voluntarios_ativos,
                         horas_por_mes=horas_por_mes,
                         ano_atual=hoje.year)


@app.route('/admin/usuarios')
@login_required
@admin_required
def admin_usuarios():
    voluntarios = Voluntario.query.order_by(Voluntario.nome).all()
    
    usuarios_stats = []
    for vol in voluntarios:
        total_horas = db.session.query(func.sum(Diario.duracao)).filter(
            Diario.voluntario_id == vol.id
        ).scalar() or 0
        
        total_atuacoes = Diario.query.filter_by(voluntario_id=vol.id).count()
        
        ultima_atuacao = Diario.query.filter_by(voluntario_id=vol.id).order_by(
            Diario.data.desc()
        ).first()
        
        usuarios_stats.append({
            'voluntario': vol,
            'total_horas': total_horas,
            'total_atuacoes': total_atuacoes,
            'ultima_atuacao': ultima_atuacao.data if ultima_atuacao else None
        })
    
    return render_template('admin_usuarios.html', usuarios_stats=usuarios_stats)


@app.route('/admin/usuarios/criar', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_criar_usuario():
    if request.method == 'POST':
        nome = request.form.get('nome')
        email = request.form.get('email')
        senha = request.form.get('senha')
        estado_padrao = request.form.get('estado_padrao')
        hospital_padrao_id = request.form.get('hospital_padrao_id')
        is_admin = request.form.get('is_admin') == 'on'
        
        if not nome or not email or not senha:
            flash('Nome, email e senha são obrigatórios.', 'error')
            return redirect(url_for('admin_criar_usuario'))
        
        existente = Voluntario.query.filter_by(email=email).first()
        if existente:
            flash('Já existe um usuário com este e-mail.', 'error')
            return redirect(url_for('admin_criar_usuario'))
        
        validated_hospital_id = None
        if hospital_padrao_id:
            hospital = Hospital.query.get(int(hospital_padrao_id))
            if hospital:
                validated_hospital_id = hospital.id
        
        novo_voluntario = Voluntario(
            nome=nome,
            email=email,
            estado_padrao=estado_padrao if estado_padrao else None,
            hospital_padrao_id=validated_hospital_id,
            is_admin=is_admin
        )
        novo_voluntario.set_senha(senha)
        
        db.session.add(novo_voluntario)
        db.session.commit()
        
        flash(f'Usuário {nome} criado com sucesso!', 'success')
        return redirect(url_for('admin_usuarios'))
    
    hospitais = Hospital.query.order_by(Hospital.estado, Hospital.nome).all()
    estados = db.session.query(Hospital.estado).distinct().order_by(Hospital.estado).all()
    estados = [e[0] for e in estados]
    
    return render_template('admin_criar_usuario.html', hospitais=hospitais, estados=estados)


@app.route('/admin/usuarios/excluir/<int:usuario_id>', methods=['POST'])
@login_required
@admin_required
def admin_excluir_usuario(usuario_id):
    voluntario = Voluntario.query.get_or_404(usuario_id)
    
    if voluntario.id == current_user.id:
        flash('Você não pode excluir sua própria conta.', 'error')
        return redirect(url_for('admin_usuarios'))
    
    if voluntario.is_admin:
        admin_count = Voluntario.query.filter_by(is_admin=True).count()
        if admin_count <= 1:
            flash('Não é possível excluir o último administrador do sistema.', 'error')
            return redirect(url_for('admin_usuarios'))
    
    db.session.delete(voluntario)
    db.session.commit()
    
    flash(f'Usuário {voluntario.nome} excluído com sucesso.', 'success')
    return redirect(url_for('admin_usuarios'))


@app.route('/admin/usuarios/editar/<int:usuario_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_editar_usuario(usuario_id):
    voluntario = Voluntario.query.get_or_404(usuario_id)
    
    if request.method == 'POST':
        nome = request.form.get('nome')
        email = request.form.get('email')
        nova_senha = request.form.get('nova_senha')
        estado_padrao = request.form.get('estado_padrao')
        hospital_padrao_id = request.form.get('hospital_padrao_id')
        is_admin = request.form.get('is_admin') == 'on'
        
        if not nome or not email:
            flash('Nome e email são obrigatórios.', 'error')
            return redirect(url_for('admin_editar_usuario', usuario_id=usuario_id))
        
        existente = Voluntario.query.filter(Voluntario.email == email, Voluntario.id != usuario_id).first()
        if existente:
            flash('Já existe outro usuário com este e-mail.', 'error')
            return redirect(url_for('admin_editar_usuario', usuario_id=usuario_id))
        
        if voluntario.is_admin and not is_admin:
            admin_count = Voluntario.query.filter_by(is_admin=True).count()
            if admin_count <= 1:
                flash('Não é possível remover o privilégio de administrador do último admin.', 'error')
                return redirect(url_for('admin_editar_usuario', usuario_id=usuario_id))
        
        voluntario.nome = nome
        voluntario.email = email
        voluntario.estado_padrao = estado_padrao if estado_padrao else None
        
        if hospital_padrao_id:
            hospital = Hospital.query.get(int(hospital_padrao_id))
            if hospital:
                voluntario.hospital_padrao_id = hospital.id
            else:
                voluntario.hospital_padrao_id = None
        else:
            voluntario.hospital_padrao_id = None
        
        voluntario.is_admin = is_admin
        
        if nova_senha:
            voluntario.set_senha(nova_senha)
        
        db.session.commit()
        flash(f'Usuário {nome} atualizado com sucesso!', 'success')
        return redirect(url_for('admin_usuarios'))
    
    hospitais = Hospital.query.order_by(Hospital.estado, Hospital.nome).all()
    estados = db.session.query(Hospital.estado).distinct().order_by(Hospital.estado).all()
    estados = [e[0] for e in estados]
    
    return render_template('admin_editar_usuario.html', voluntario=voluntario, hospitais=hospitais, estados=estados)


@app.route('/admin/hospitais')
@login_required
@admin_required
def admin_hospitais():
    hospitais = Hospital.query.order_by(Hospital.estado, Hospital.nome).all()
    return render_template('admin_hospitais.html', hospitais=hospitais)


@app.route('/admin/hospitais/criar', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_criar_hospital():
    if request.method == 'POST':
        nome = request.form.get('nome')
        estado = request.form.get('estado')
        
        if not nome or not estado:
            flash('Nome e estado são obrigatórios.', 'error')
            return redirect(url_for('admin_criar_hospital'))
        
        hospital = Hospital(nome=nome, estado=estado.upper())
        db.session.add(hospital)
        db.session.commit()
        
        flash(f'Hospital {nome} criado com sucesso!', 'success')
        return redirect(url_for('admin_hospitais'))
    
    estados_brasil = ['AC', 'AL', 'AP', 'AM', 'BA', 'CE', 'DF', 'ES', 'GO', 'MA', 'MT', 'MS', 'MG', 'PA', 'PB', 'PR', 'PE', 'PI', 'RJ', 'RN', 'RS', 'RO', 'RR', 'SC', 'SP', 'SE', 'TO']
    return render_template('admin_criar_hospital.html', estados=estados_brasil)


@app.route('/admin/hospitais/editar/<int:hospital_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_editar_hospital(hospital_id):
    hospital = Hospital.query.get_or_404(hospital_id)
    
    if request.method == 'POST':
        nome = request.form.get('nome')
        estado = request.form.get('estado')
        
        if not nome or not estado:
            flash('Nome e estado são obrigatórios.', 'error')
            return redirect(url_for('admin_editar_hospital', hospital_id=hospital_id))
        
        hospital.nome = nome
        hospital.estado = estado.upper()
        db.session.commit()
        
        flash(f'Hospital {nome} atualizado com sucesso!', 'success')
        return redirect(url_for('admin_hospitais'))
    
    estados_brasil = ['AC', 'AL', 'AP', 'AM', 'BA', 'CE', 'DF', 'ES', 'GO', 'MA', 'MT', 'MS', 'MG', 'PA', 'PB', 'PR', 'PE', 'PI', 'RJ', 'RN', 'RS', 'RO', 'RR', 'SC', 'SP', 'SE', 'TO']
    return render_template('admin_editar_hospital.html', hospital=hospital, estados=estados_brasil)


@app.route('/admin/hospitais/excluir/<int:hospital_id>', methods=['POST'])
@login_required
@admin_required
def admin_excluir_hospital(hospital_id):
    hospital = Hospital.query.get_or_404(hospital_id)
    
    voluntarios_vinculados = Voluntario.query.filter_by(hospital_padrao_id=hospital_id).count()
    if voluntarios_vinculados > 0:
        flash(f'Não é possível excluir. Existem {voluntarios_vinculados} voluntário(s) vinculado(s) a este hospital.', 'error')
        return redirect(url_for('admin_hospitais'))
    
    db.session.delete(hospital)
    db.session.commit()
    
    flash(f'Hospital {hospital.nome} excluído com sucesso.', 'success')
    return redirect(url_for('admin_hospitais'))


@app.route('/admin/livros')
@login_required
@admin_required
def admin_livros():
    livros = Livro.query.order_by(Livro.titulo).all()
    return render_template('admin_livros.html', livros=livros)


@app.route('/admin/livros/criar', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_criar_livro():
    if request.method == 'POST':
        titulo = request.form.get('titulo')
        autor = request.form.get('autor')
        editora = request.form.get('editora')
        
        if not titulo:
            flash('Título é obrigatório.', 'error')
            return redirect(url_for('admin_criar_livro'))
        
        livro = Livro(titulo=titulo, autor=autor, editora=editora)
        db.session.add(livro)
        db.session.commit()
        
        flash(f'Livro "{titulo}" adicionado com sucesso!', 'success')
        return redirect(url_for('admin_livros'))
    
    return render_template('admin_criar_livro.html')


@app.route('/admin/livros/editar/<int:livro_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_editar_livro(livro_id):
    livro = Livro.query.get_or_404(livro_id)
    
    if request.method == 'POST':
        titulo = request.form.get('titulo')
        autor = request.form.get('autor')
        editora = request.form.get('editora')
        
        if not titulo:
            flash('Título é obrigatório.', 'error')
            return redirect(url_for('admin_editar_livro', livro_id=livro_id))
        
        livro.titulo = titulo
        livro.autor = autor
        livro.editora = editora
        db.session.commit()
        
        flash(f'Livro "{titulo}" atualizado com sucesso!', 'success')
        return redirect(url_for('admin_livros'))
    
    return render_template('admin_editar_livro.html', livro=livro)


@app.route('/admin/livros/excluir/<int:livro_id>', methods=['POST'])
@login_required
@admin_required
def admin_excluir_livro(livro_id):
    livro = Livro.query.get_or_404(livro_id)
    
    db.session.delete(livro)
    db.session.commit()
    
    flash(f'Livro "{livro.titulo}" excluído com sucesso.', 'success')
    return redirect(url_for('admin_livros'))


@app.route('/admin/relatorios/exportar')
@login_required
@admin_required
def admin_exportar_relatorio():
    import csv
    from io import StringIO
    from flask import Response
    
    periodo = request.args.get('periodo', 'mes')
    voluntario_id = request.args.get('voluntario_id', '')
    
    hoje = datetime.now()
    
    if periodo == 'mes':
        inicio = hoje.replace(day=1)
    elif periodo == 'ano':
        inicio = hoje.replace(month=1, day=1)
    else:
        inicio = datetime(2000, 1, 1)
    
    query = Diario.query.filter(Diario.data >= inicio)
    
    if voluntario_id:
        query = query.filter(Diario.voluntario_id == int(voluntario_id))
    
    diarios = query.order_by(Diario.data.desc()).all()
    
    output = StringIO()
    writer = csv.writer(output)
    
    writer.writerow(['Data', 'Voluntário', 'Período', 'Duração (h)', 'Total Pacientes', 'Livros Contados', 'Locais'])
    
    for diario in diarios:
        total_pacientes = 0
        if diario.pacientes_atendidos:
            for faixa, dados in diario.pacientes_atendidos.items():
                total_pacientes += dados.get('feminino', 0) + dados.get('masculino', 0)
        
        livros = len(diario.livros_contados) if diario.livros_contados else 0
        locais = ', '.join(diario.locais_atendimento) if diario.locais_atendimento else ''
        
        writer.writerow([
            diario.data.strftime('%d/%m/%Y'),
            diario.voluntario.nome,
            diario.periodo,
            f'{diario.duracao:.1f}',
            total_pacientes,
            livros,
            locais
        ])
    
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=relatorio_atuacoes_{hoje.strftime("%Y%m%d")}.csv'}
    )


@app.route('/admin/relatorios/exportar-xlsx')
@login_required
@admin_required
def admin_exportar_xlsx():
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import Response
    
    periodo = request.args.get('periodo', 'mes')
    voluntario_id = request.args.get('voluntario_id', '')
    
    hoje = datetime.now()
    
    if periodo == 'mes':
        inicio = hoje.replace(day=1)
        titulo_periodo = f'{inicio.strftime("%B de %Y")}'
    elif periodo == 'ano':
        inicio = hoje.replace(month=1, day=1)
        titulo_periodo = f'Ano {hoje.year}'
    else:
        inicio = datetime(2000, 1, 1)
        titulo_periodo = 'Todo o período'
    
    query = Diario.query.filter(Diario.data >= inicio)
    
    if voluntario_id:
        query = query.filter(Diario.voluntario_id == int(voluntario_id))
    
    diarios = query.order_by(Diario.data.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Atuações"
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:G1')
    ws['A1'] = f"Relatório de Atuações - {titulo_periodo}"
    ws['A1'].font = Font(bold=True, size=16, color="2E8B57")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Gerado em: {hoje.strftime('%d/%m/%Y às %H:%M')}"
    ws['A2'].alignment = Alignment(horizontal="center")
    
    total_horas = sum(d.duracao for d in diarios)
    total_atuacoes = len(diarios)
    total_pacientes = 0
    for diario in diarios:
        if diario.pacientes_atendidos:
            for faixa, dados in diario.pacientes_atendidos.items():
                total_pacientes += dados.get('feminino', 0) + dados.get('masculino', 0)
    
    ws['A4'] = "Total de Horas:"
    ws['B4'] = f"{total_horas:.1f}h"
    ws['C4'] = "Total de Atuações:"
    ws['D4'] = total_atuacoes
    ws['E4'] = "Total de Pacientes:"
    ws['F4'] = total_pacientes
    for col in ['A', 'C', 'E']:
        ws[f'{col}4'].font = Font(bold=True)
    
    headers = ['Data', 'Voluntário', 'Período', 'Duração (h)', 'Pacientes', 'Livros', 'Locais']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row, diario in enumerate(diarios, 7):
        total_pac = 0
        if diario.pacientes_atendidos:
            for faixa, dados in diario.pacientes_atendidos.items():
                total_pac += dados.get('feminino', 0) + dados.get('masculino', 0)
        
        livros = len(diario.livros_contados) if diario.livros_contados else 0
        locais = ', '.join(diario.locais_atendimento) if diario.locais_atendimento else ''
        
        data_row = [
            diario.data.strftime('%d/%m/%Y'),
            diario.voluntario.nome,
            diario.periodo,
            diario.duracao,
            total_pac,
            livros,
            locais
        ]
        
        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if col == 4:
                cell.number_format = '0.0'
    
    column_widths = [12, 25, 12, 12, 12, 10, 35]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=relatorio_atuacoes_{hoje.strftime("%Y%m%d")}.xlsx'}
    )


@app.route('/admin/relatorios/exportar-pdf')
@login_required
@admin_required
def admin_exportar_pdf():
    from io import BytesIO
    from flask import Response
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.piecharts import Pie
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    
    periodo = request.args.get('periodo', 'mes')
    voluntario_id = request.args.get('voluntario_id', '')
    
    hoje = datetime.now()
    
    if periodo == 'mes':
        inicio = hoje.replace(day=1)
        titulo_periodo = f'{inicio.strftime("%B de %Y")}'
    elif periodo == 'ano':
        inicio = hoje.replace(month=1, day=1)
        titulo_periodo = f'Ano {hoje.year}'
    else:
        inicio = datetime(2000, 1, 1)
        titulo_periodo = 'Todo o período'
    
    query = Diario.query.filter(Diario.data >= inicio)
    
    if voluntario_id:
        query = query.filter(Diario.voluntario_id == int(voluntario_id))
    
    diarios = query.order_by(Diario.data.desc()).all()
    
    total_horas = sum(d.duracao for d in diarios)
    total_atuacoes = len(diarios)
    
    total_pacientes = 0
    pacientes_por_faixa = {}
    for diario in diarios:
        if diario.pacientes_atendidos:
            for faixa, dados in diario.pacientes_atendidos.items():
                fem = dados.get('feminino', 0)
                masc = dados.get('masculino', 0)
                total_pacientes += fem + masc
                if faixa not in pacientes_por_faixa:
                    pacientes_por_faixa[faixa] = {'feminino': 0, 'masculino': 0}
                pacientes_por_faixa[faixa]['feminino'] += fem
                pacientes_por_faixa[faixa]['masculino'] += masc
    
    locais_contagem = {}
    for diario in diarios:
        if diario.locais_atendimento:
            for local in diario.locais_atendimento:
                locais_contagem[local] = locais_contagem.get(local, 0) + 1
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#2E8B57'),
        spaceAfter=12,
        alignment=1
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.gray,
        spaceAfter=20,
        alignment=1
    )
    
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#2E8B57'),
        spaceBefore=20,
        spaceAfter=10
    )
    
    elements.append(Paragraph("Diário do Contador", title_style))
    elements.append(Paragraph(f"Relatório de Atuações - {titulo_periodo}", subtitle_style))
    elements.append(Paragraph(f"Gerado em: {hoje.strftime('%d/%m/%Y às %H:%M')}", subtitle_style))
    
    elements.append(Paragraph("Resumo Geral", section_style))
    
    summary_data = [
        ['Indicador', 'Valor'],
        ['Total de Horas', f'{total_horas:.1f}h'],
        ['Total de Atuações', str(total_atuacoes)],
        ['Total de Pacientes Atendidos', str(total_pacientes)]
    ]
    
    summary_table = Table(summary_data, colWidths=[10*cm, 5*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E8B57')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0FFF0')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#2E8B57')),
        ('FONTSIZE', (0, 1), (-1, -1), 11),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    if pacientes_por_faixa:
        elements.append(Paragraph("Pacientes por Faixa Etária", section_style))
        
        faixas = ['0-3', '4-6', '7-9', '10-12', '13-15', '16-18']
        faixa_data = [['Faixa Etária', 'Feminino', 'Masculino', 'Total']]
        
        for faixa in faixas:
            if faixa in pacientes_por_faixa:
                fem = pacientes_por_faixa[faixa]['feminino']
                masc = pacientes_por_faixa[faixa]['masculino']
                faixa_data.append([f'{faixa} anos', str(fem), str(masc), str(fem + masc)])
        
        if len(faixa_data) > 1:
            faixa_table = Table(faixa_data, colWidths=[5*cm, 3*cm, 3*cm, 3*cm])
            faixa_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF6B6B')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#FF6B6B')),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FFF0F0')),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(faixa_table)
            elements.append(Spacer(1, 20))
            
            fig, ax = plt.subplots(figsize=(6, 3))
            faixas_labels = [f'{f} anos' for f in faixas if f in pacientes_por_faixa]
            fem_values = [pacientes_por_faixa.get(f, {}).get('feminino', 0) for f in faixas if f in pacientes_por_faixa]
            masc_values = [pacientes_por_faixa.get(f, {}).get('masculino', 0) for f in faixas if f in pacientes_por_faixa]
            
            x = range(len(faixas_labels))
            width = 0.35
            ax.bar([i - width/2 for i in x], fem_values, width, label='Feminino', color='#FF6B6B')
            ax.bar([i + width/2 for i in x], masc_values, width, label='Masculino', color='#4ECDC4')
            ax.set_ylabel('Pacientes')
            ax.set_title('Pacientes por Faixa Etária e Gênero')
            ax.set_xticks(x)
            ax.set_xticklabels(faixas_labels, rotation=45, ha='right')
            ax.legend()
            plt.tight_layout()
            
            chart_buffer = BytesIO()
            plt.savefig(chart_buffer, format='png', dpi=150, bbox_inches='tight')
            plt.close()
            chart_buffer.seek(0)
            
            chart_img = Image(chart_buffer, width=14*cm, height=7*cm)
            elements.append(chart_img)
            elements.append(Spacer(1, 20))
    
    if locais_contagem:
        elements.append(Paragraph("Locais de Atendimento", section_style))
        
        fig, ax = plt.subplots(figsize=(5, 4))
        locais = list(locais_contagem.keys())
        valores = list(locais_contagem.values())
        colors_pie = ['#2E8B57', '#FF6B6B', '#FFD93D', '#4ECDC4', '#9B59B6', '#3498DB', '#E67E22', '#1ABC9C']
        
        ax.pie(valores, labels=locais, autopct='%1.1f%%', colors=colors_pie[:len(locais)], startangle=90)
        ax.set_title('Distribuição por Local de Atendimento')
        
        pie_buffer = BytesIO()
        plt.savefig(pie_buffer, format='png', dpi=150, bbox_inches='tight')
        plt.close()
        pie_buffer.seek(0)
        
        pie_img = Image(pie_buffer, width=12*cm, height=10*cm)
        elements.append(pie_img)
        elements.append(Spacer(1, 20))
    
    if diarios:
        elements.append(Paragraph("Atuações Detalhadas", section_style))
        
        detail_data = [['Data', 'Voluntário', 'Período', 'Horas', 'Pacientes']]
        
        for diario in diarios[:50]:
            total_pac = 0
            if diario.pacientes_atendidos:
                for faixa, dados in diario.pacientes_atendidos.items():
                    total_pac += dados.get('feminino', 0) + dados.get('masculino', 0)
            
            detail_data.append([
                diario.data.strftime('%d/%m/%Y'),
                diario.voluntario.nome[:20],
                diario.periodo,
                f'{diario.duracao:.1f}h',
                str(total_pac)
            ])
        
        detail_table = Table(detail_data, colWidths=[2.5*cm, 5*cm, 2.5*cm, 2*cm, 2.5*cm])
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E8B57')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0FFF0')]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(detail_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename=relatorio_atuacoes_{hoje.strftime("%Y%m%d")}.pdf'}
    )


@app.route('/admin/relatorios')
@login_required
@admin_required
def admin_relatorios():
    periodo = request.args.get('periodo', 'mes')
    voluntario_id = request.args.get('voluntario_id', '')
    
    hoje = datetime.now()
    
    if periodo == 'mes':
        inicio = hoje.replace(day=1)
        titulo_periodo = f'{inicio.strftime("%B de %Y")}'
    elif periodo == 'ano':
        inicio = hoje.replace(month=1, day=1)
        titulo_periodo = f'Ano {hoje.year}'
    else:
        inicio = datetime(2000, 1, 1)
        titulo_periodo = 'Todo o período'
    
    query = Diario.query.filter(Diario.data >= inicio)
    
    if voluntario_id:
        query = query.filter(Diario.voluntario_id == int(voluntario_id))
    
    diarios = query.order_by(Diario.data.desc()).all()
    
    total_horas = sum(d.duracao for d in diarios)
    total_atuacoes = len(diarios)
    
    total_pacientes = 0
    pacientes_por_faixa = {}
    for diario in diarios:
        if diario.pacientes_atendidos:
            for faixa, dados in diario.pacientes_atendidos.items():
                fem = dados.get('feminino', 0)
                masc = dados.get('masculino', 0)
                total_pacientes += fem + masc
                
                if faixa not in pacientes_por_faixa:
                    pacientes_por_faixa[faixa] = {'feminino': 0, 'masculino': 0}
                pacientes_por_faixa[faixa]['feminino'] += fem
                pacientes_por_faixa[faixa]['masculino'] += masc
    
    livros_contados = {}
    for diario in diarios:
        if diario.livros_contados:
            for livro in diario.livros_contados:
                titulo = livro.get('titulo', 'Desconhecido')
                if titulo not in livros_contados:
                    livros_contados[titulo] = {
                        'titulo': titulo,
                        'autor': livro.get('autor', ''),
                        'editora': livro.get('editora', ''),
                        'vezes': 0
                    }
                livros_contados[titulo]['vezes'] += 1
    
    livros_ranking = sorted(livros_contados.values(), key=lambda x: x['vezes'], reverse=True)[:10]
    
    locais_contagem = {}
    for diario in diarios:
        if diario.locais_atendimento:
            for local in diario.locais_atendimento:
                locais_contagem[local] = locais_contagem.get(local, 0) + 1
    
    voluntarios = Voluntario.query.order_by(Voluntario.nome).all()
    
    return render_template('admin_relatorios.html',
                         diarios=diarios,
                         total_horas=total_horas,
                         total_atuacoes=total_atuacoes,
                         total_pacientes=total_pacientes,
                         pacientes_por_faixa=pacientes_por_faixa,
                         livros_ranking=livros_ranking,
                         locais_contagem=locais_contagem,
                         periodo=periodo,
                         titulo_periodo=titulo_periodo,
                         voluntarios=voluntarios,
                         voluntario_selecionado=voluntario_id)
